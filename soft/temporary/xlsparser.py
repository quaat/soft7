from pydantic.dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from itertools import product
import numpy as np
import re
from soft.temporary.xlsconfig import Workbook, WorkbookField


allowed_types = {
    'string': str,
    'float': float,
    'complex': complex,
    'int': int,
    'dict': dict,
    'boolean': bool,
    'bytes': bytes,
    'bytearray': bytearray
}


def validate_data(tup):
    rexp, typ, val = tup
    if val == None or not isinstance(val, typ):
        return False
    if rexp:
        val = bool(re.match(rexp, val))
        return val
    return True


@dataclass
class XLSParser:
    """ Parse excel documents given document configuration
        and an S7-entity
    """

    xls_model: Workbook

    def get(self):
        wb_obj = load_workbook(self.xls_model.workbook.path, data_only=True)
        sheet_obj = wb_obj[self.xls_model.sheet]

        alldata = {}
        redata = {}
        typedata = {}
        for d in self.xls_model.data:
            idx = d[0]
            prop: WorkbookField = d[1]

            item = []
            dataRange = CellRange(str(self.xls_model.data_range))
            cellRange = dataRange.intersection(CellRange(prop.range))
            for cell in product(
                range(cellRange.min_row, cellRange.max_row+1),
                range(cellRange.min_col, cellRange.max_col+1)
            ):
                cell_obj = sheet_obj.cell(*cell)
                item.append(cell_obj.value)
            alldata[idx] = item
            redata[idx] = prop.regexp
            typedata[idx] = allowed_types[prop.type]

        newdata = {}
        for elem in np.column_stack(
            [alldata[label] for label in alldata.keys()]
        ):
            tuples = list(zip([redata[label] for label in alldata.keys()],
                         [typedata[label] for label in alldata.keys()],
                         elem))
            if not False in [validate_data(tup) for tup in tuples]:
                for (k,v) in (list(zip(alldata.keys(), elem))):
                    if not k in newdata:
                        newdata[k] = []
                    newdata[k].append(v)
        return newdata
